"""Build data/tag_sites/tedman_gpcr_controls.tsv from the Tedman GPCR screen.

Join key: `Receptor Name` (SYMBOL_ENST) between the surface screen and the
plasmid-map index. Canonical UniProt via resolve_by_hgnc_id (HGNC-ID path — never
bare symbols); canonical sequence via _uniprot_entry. Junction projected to
canonical numbering + verified against the sequence.

    uv run python scripts/build_tedman_gpcr_controls.py            # writes the TSV
    uv run python scripts/build_tedman_gpcr_controls.py --limit 20 # smoke subset
"""
from __future__ import annotations

import argparse
import csv
import functools
import logging
from pathlib import Path

import openpyxl
import pandas as pd

from accessible_surfaceome.env import load_env
from accessible_surfaceome.paths import REPO_ROOT
from accessible_surfaceome.tag_sites.tedman import map_junction_to_canonical, parse_ha_position
from accessible_surfaceome.tools._shared.http import open_default_client
from accessible_surfaceome.tools.gene_lookup import _uniprot_entry, resolve_by_hgnc_id

log = logging.getLogger("build_tedman")
SRC = REPO_ROOT / "data" / "external" / "tedman_gpcr_screen"
OUT = REPO_ROOT / "data" / "tag_sites" / "tedman_gpcr_controls.tsv"
COHORT = REPO_ROOT / "data" / "external" / "ncbi_gene_info" / "Homo_sapiens.protein_coding.with_hgnc.tsv"
COLS = [
    "gene_symbol", "hgnc_id", "uniprot_acc", "ensembl_gene_id", "ensembl_transcript_id",
    "is_canonical", "gpcr_class", "site_kind", "junction_after_residue", "expected_residue",
    "tag", "tag_length", "surface_expression_pme", "surface_expression_sd",
    "verified", "source_key",
]


@functools.lru_cache(maxsize=1)
def _cohort() -> pd.DataFrame:
    return pd.read_csv(COHORT, sep="\t", dtype=str)


@functools.lru_cache(maxsize=1)
def _symbol_to_hgnc() -> dict[str, str]:
    df = _cohort()
    return {s: h for s, h in zip(df["gene_symbol"], df["hgnc_id"]) if isinstance(h, str) and h.strip()}


@functools.lru_cache(maxsize=1)
def _ensg_to_hgnc() -> dict[str, str]:
    df = _cohort()
    return {e: h for e, h in zip(df["ensembl_gene"], df["hgnc_id"]) if isinstance(h, str) and h.strip()}


def _resolve(symbol: str, ensg: str, http) -> tuple[str, str, str]:
    """(hgnc_id, canonical uniprot_acc, canonical sequence). HGNC-ID path only."""
    hgnc_id = _symbol_to_hgnc().get(symbol) or (_ensg_to_hgnc().get(ensg) if ensg else None)
    if not hgnc_id:
        raise ValueError(f"no hgnc_id for {symbol} / {ensg}")
    bundle = resolve_by_hgnc_id(hgnc_id, http=http)
    entry = _uniprot_entry(bundle.uniprot_acc, http=http)
    seq = (entry.get("sequence") or {}).get("value") or ""
    return hgnc_id, bundle.uniprot_acc, seq


def _index_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    idx = {name: i for i, name in enumerate(hdr)}

    def g(r, name):
        for k, i in idx.items():
            if k and k.lower().startswith(name.lower()):
                return r[i]
        return None

    out = []
    for r in rows[1:]:
        if not g(r, "hgnc_symbol"):
            continue
        out.append({
            "symbol": g(r, "hgnc_symbol"), "uniprot": g(r, "uniprot_id"),
            "ensg": g(r, "ensembl_gene_id"), "enst": g(r, "ensembl_transcript_id"),
            "gpcr_class": g(r, "gpcr_class"), "alt": g(r, "alt_isoform"),
            "ha": g(r, "ha_insert_position"),
        })
    return out


def _staining(path: Path) -> dict[str, tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Canonical"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    ix = {n: i for i, n in enumerate(hdr)}
    out = {}
    for r in rows[1:]:
        name = r[ix["Receptor Name"]]
        if not name:
            continue
        key = "_".join(str(name).split("_")[:2])   # SYMBOL_ENST
        out[key] = (r[ix["Immunostaining Intensity"]], r[ix["Immunostaining Standard Deviation"]])
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    load_env()
    http = open_default_client()
    index = _index_rows(SRC / "GPCR_Library_Index_Final.xlsx")
    stain = _staining(SRC / "tedman2025_gpcr_screen_media-2.xlsx")
    if args.limit:
        index = index[: args.limit]

    seq_cache: dict[str, tuple[str, str, str]] = {}
    written = 0
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLS, delimiter="\t")
        w.writeheader()
        for row in index:
            sym = row["symbol"]
            if sym not in seq_cache:
                try:
                    seq_cache[sym] = _resolve(sym, row["ensg"], http)
                except Exception as e:  # noqa: BLE001
                    log.warning("resolve failed %s: %s", sym, e)
                    seq_cache[sym] = ("", "", "")
            hgnc_id, acc, seq = seq_cache[sym]
            jm = map_junction_to_canonical(parse_ha_position(row["ha"]), seq)
            pme, sd = stain.get(f"{sym}_{row['enst']}", (None, None))
            w.writerow({
                "gene_symbol": sym, "hgnc_id": hgnc_id, "uniprot_acc": acc,
                "ensembl_gene_id": row["ensg"], "ensembl_transcript_id": row["enst"],
                "is_canonical": str(row["alt"] == "canonical").lower(),
                "gpcr_class": row["gpcr_class"], "site_kind": "terminal_n",
                "junction_after_residue": "" if jm.insert_after_residue is None else jm.insert_after_residue,
                "expected_residue": jm.residue_before or (jm.residue_after or ""),
                "tag": "HA", "tag_length": 9,
                "surface_expression_pme": "" if pme is None else round(float(pme), 1),
                "surface_expression_sd": "" if sd is None else round(float(sd), 1),
                "verified": str(jm.verified and bool(acc)).lower(), "source_key": "tedman2026",
            })
            written += 1
    log.info("wrote %d rows -> %s", written, OUT)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
